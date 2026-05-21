import re
import unicodedata
from datetime import date, datetime
from io import BytesIO

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.auth import get_db, get_usuario_logado, somente_gestor, checar_acesso_condominio
from app.models.despesa import Despesa
from app.models.morador import Morador
from app.models.receita import Receita
router = APIRouter(prefix="/importar", tags=["Importação"])


# ── Helpers ───────────────────────────────────────────────────

def _norm(s: str) -> str:
    """Normaliza string: remove acentos, lowercase, só alfanumérico."""
    nfkd = unicodedata.normalize("NFKD", str(s))
    ascii_ = nfkd.encode("ASCII", "ignore").decode()
    return re.sub(r"[^a-z0-9]", "", ascii_.lower())


# Aliases aceitos por coluna em cada tipo de planilha
_ALIASES = {
    "despesas": {
        "descricao": {"descricao", "desc", "description", "historico", "lancamento", "item", "nome"},
        "valor":     {"valor", "value", "vl", "preco", "montante", "custo"},
        "data":      {"data", "date", "dt", "vencimento", "competencia", "periodo"},
    },
    "receitas": {
        "descricao": {"descricao", "desc", "description", "historico", "lancamento", "item", "nome"},
        "valor":     {"valor", "value", "vl", "preco", "montante"},
        "data":      {"data", "date", "dt", "recebimento", "competencia"},
    },
    "moradores": {
        "nome":       {"nome", "name", "morador", "residente", "proprietario"},
        "email":      {"email", "mail", "correio", "emailaddr"},
        "telefone":   {"telefone", "fone", "phone", "tel", "celular", "contato", "whatsapp"},
        "apartamento":{"apartamento", "apto", "apt", "unidade", "ap", "sala", "bloco", "numero"},
    },
}

_TEMPLATES = {
    "despesas":  [["descricao", "valor", "data"],
                  ["Manutenção elevador", "500.00", "15/01/2025"],
                  ["Limpeza área comum", "300.00", "20/01/2025"]],
    "receitas":  [["descricao", "valor", "data"],
                  ["Taxa condominial - Apto 101", "350.00", "05/01/2025"],
                  ["Taxa condominial - Apto 202", "350.00", "05/01/2025"]],
    "moradores": [["nome", "email", "telefone", "apartamento"],
                  ["João Silva", "joao@email.com", "11999990000", "101"],
                  ["Maria Santos", "maria@email.com", "11988880000", "202"]],
}


def _mapear_colunas(headers: list, aliases: dict) -> dict:
    """Retorna {campo_canonical: índice_coluna}."""
    mapa = {}
    for i, h in enumerate(headers):
        norm = _norm(str(h))
        for campo, alias_set in aliases.items():
            if norm in alias_set and campo not in mapa:
                mapa[campo] = i
    return mapa


def _parse_data(val) -> date:
    if val is None:
        return date.today()
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d",
                "%d/%m/%y", "%m/%y", "%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return date.today()


def _parse_valor(val) -> float:
    if val is None:
        raise ValueError("valor vazio")
    if isinstance(val, (int, float)):
        v = float(val)
    else:
        s = re.sub(r"[R$\s]", "", str(val).strip())
        # Detecta formato brasileiro (1.234,56) vs americano (1,234.56)
        if "," in s and "." in s:
            if s.rindex(",") > s.rindex("."):   # 1.234,56
                s = s.replace(".", "").replace(",", ".")
            else:                                 # 1,234.56
                s = s.replace(",", "")
        elif "," in s:
            s = s.replace(",", ".")
        v = float(s)
    if v <= 0:
        raise ValueError("valor deve ser positivo")
    return round(v, 2)



_MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10 MB


def _ler_workbook(file: UploadFile) -> openpyxl.Workbook:
    conteudo = file.file.read(_MAX_UPLOAD_BYTES + 1)
    if len(conteudo) > _MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="Arquivo muito grande. Limite: 10 MB.")
    try:
        return openpyxl.load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Arquivo inválido. Envie um .xlsx válido.")


# ── Template download ─────────────────────────────────────────

@router.get("/template/{tipo}")
def baixar_template(tipo: str, usuario=Depends(get_usuario_logado)):
    if tipo not in _TEMPLATES:
        raise HTTPException(status_code=400, detail="Tipo inválido. Use: despesas, receitas ou moradores")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tipo.capitalize()
    for row in _TEMPLATES[tipo]:
        ws.append(row)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=template_{tipo}.xlsx"},
    )


# ── Importar Despesas ─────────────────────────────────────────

@router.post("/despesas/{condominio_id}")
def importar_despesas(
    condominio_id: int,
    arquivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    usuario=Depends(somente_gestor),
):
    checar_acesso_condominio(usuario, condominio_id)
    wb = _ler_workbook(arquivo)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Planilha vazia.")

    mapa = _mapear_colunas([str(h) if h is not None else "" for h in rows[0]],
                           _ALIASES["despesas"])

    campos_obrigatorios = {"descricao", "valor"}
    faltando = campos_obrigatorios - set(mapa)
    if faltando:
        raise HTTPException(
            status_code=400,
            detail=f"Colunas obrigatórias não encontradas: {', '.join(faltando)}. "
                   f"Baixe o template em GET /importar/template/despesas",
        )

    criados, erros = 0, []
    for n, row in enumerate(rows[1:], start=2):
        try:
            descricao = str(row[mapa["descricao"]] or "").strip()
            if not descricao:
                erros.append(f"Linha {n}: descrição vazia")
                continue
            valor = _parse_valor(row[mapa["valor"]])
            data_d = _parse_data(row[mapa["data"]] if "data" in mapa else None)

            db.add(Despesa(
                descricao=descricao,
                valor=valor,
                data=data_d,
                condominio_id=condominio_id,
            ))
            criados += 1
        except Exception as e:
            erros.append(f"Linha {n}: {e}")

    db.commit()
    return {"criados": criados, "erros": erros, "total_linhas": len(rows) - 1}


# ── Importar Receitas ─────────────────────────────────────────

@router.post("/receitas/{condominio_id}")
def importar_receitas(
    condominio_id: int,
    arquivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    usuario=Depends(somente_gestor),
):
    checar_acesso_condominio(usuario, condominio_id)
    wb = _ler_workbook(arquivo)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Planilha vazia.")

    mapa = _mapear_colunas([str(h) if h is not None else "" for h in rows[0]],
                           _ALIASES["receitas"])

    if "descricao" not in mapa or "valor" not in mapa:
        raise HTTPException(
            status_code=400,
            detail="Colunas obrigatórias não encontradas: descricao, valor. "
                   "Baixe o template em GET /importar/template/receitas",
        )

    criados, erros = 0, []
    for n, row in enumerate(rows[1:], start=2):
        try:
            descricao = str(row[mapa["descricao"]] or "").strip()
            if not descricao:
                erros.append(f"Linha {n}: descrição vazia")
                continue
            valor = _parse_valor(row[mapa["valor"]])
            data_r = _parse_data(row[mapa["data"]] if "data" in mapa else None)

            db.add(Receita(
                descricao=descricao,
                valor=valor,
                data=data_r,
                condominio_id=condominio_id,
            ))
            criados += 1
        except Exception as e:
            erros.append(f"Linha {n}: {e}")

    db.commit()
    return {"criados": criados, "erros": erros, "total_linhas": len(rows) - 1}


# ── Importar Moradores ────────────────────────────────────────

@router.post("/moradores/{condominio_id}")
def importar_moradores(
    condominio_id: int,
    arquivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    usuario=Depends(somente_gestor),
):
    checar_acesso_condominio(usuario, condominio_id)
    wb = _ler_workbook(arquivo)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Planilha vazia.")

    mapa = _mapear_colunas([str(h) if h is not None else "" for h in rows[0]],
                           _ALIASES["moradores"])

    if "nome" not in mapa:
        raise HTTPException(
            status_code=400,
            detail="Coluna 'nome' não encontrada. "
                   "Baixe o template em GET /importar/template/moradores",
        )

    criados, erros = 0, []
    for n, row in enumerate(rows[1:], start=2):
        try:
            nome = str(row[mapa["nome"]] or "").strip()
            if not nome:
                erros.append(f"Linha {n}: nome vazio")
                continue

            email = str(row[mapa["email"]] if "email" in mapa else "").strip() or None
            telefone = str(row[mapa["telefone"]] if "telefone" in mapa else "").strip() or None
            apartamento = str(row[mapa["apartamento"]] if "apartamento" in mapa else "").strip() or None

            # Verifica email duplicado
            if email:
                existente = db.query(Morador).filter(Morador.email == email).first()
                if existente:
                    erros.append(f"Linha {n}: e-mail '{email}' já cadastrado")
                    continue

            db.add(Morador(
                nome=nome,
                email=email,
                telefone=telefone,
                apartamento=apartamento,
                condominio_id=condominio_id,
            ))
            criados += 1
        except Exception as e:
            erros.append(f"Linha {n}: {e}")

    db.commit()
    return {"criados": criados, "erros": erros, "total_linhas": len(rows) - 1}
